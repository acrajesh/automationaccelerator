import os

import re

import pandas as pd

from openpyxl import Workbook

from openpyxl.utils.dataframe import dataframe_to_rows

def resolve_copybooks(cobol_source: str, copybook_dir: str) -> str:
    resolved_lines = []
    copybook_pattern = re.compile(r"^\s*COPY\s+['\"]?(\S+)['\"]?", re.IGNORECASE)
    lines = cobol_source.splitlines()
    
    for line in lines:
        match = copybook_pattern.search(line)
        
        if match:
            copybook_name = match.group(1).strip().upper()
            copybook_file = None
            
            for ext in ['', '.cpy', '.CPY', '.cob', '.COB', '.cbl', '.CBL']:
                path = os.path.join(copybook_dir, copybook_name.lower() + ext)
                if os.path.exists(path):
                    copybook_file = path
                    break
            
            if copybook_file:
                try:
                    # Try UTF-8 first
                    with open(copybook_file, 'r', encoding='utf-8', errors='replace') as cbf:
                        copybook_content = cbf.read()
                except UnicodeDecodeError:
                    # Fall back to Latin-1 which can read any byte
                    with open(copybook_file, 'r', encoding='latin-1') as cbf:
                        copybook_content = cbf.read()
                
                resolved_lines.append(f"*> Resolved: {line.strip()}")
                resolved_lines.extend(copybook_content.splitlines())
            else:
                resolved_lines.append(f"*> COPYBOOK NOT FOUND: {line.strip()}")
        else:
            resolved_lines.append(line)
    
    return '\n'.join(resolved_lines)

def detect_ims_in_cobol(file_content: str, program_name: str):
    """Detect IMS calls in COBOL source code using universal patterns.
    
    Looks for standard IMS entry points and function codes to identify true IMS calls.
    """
    ims_calls = []
    param_counts = {}
    function_counts = {}
    
    lines = file_content.splitlines()
    n = len(lines)
    i = 0
    
    # IMS entry point patterns - these are the standard entry points for IMS calls
    ims_entry_points = [
        r'\bCBLTDLI\b', r'\bAERTDLI\b', r'\bPLITDLI\b', 
        r'\bDFSLI000\b', r'\bDFHEI1\b', r'\bDLITDLI\b'
    ]
    
    # IMS function codes - these are the standard IMS function codes
    ims_functions = [
        'GU', 'GN', 'GNP', 'GHU', 'GHN', 'GHNP',
        'ISRT', 'DLET', 'REPL', 'CHKP', 'XRST', 'ROLL',
        'ROLS', 'SETS', 'SETU', 'INIT', 'INQY', 'CMD',
        'GCMD', 'AUTH', 'QUERY', 'POS', 'PCB'
    ]
    
    while i < n:
        line = lines[i].strip().upper()
        
        # Check for CALL statement with IMS entry points
        if "CALL" in line:
            # Check for direct IMS entry point calls - this is the most reliable way to detect IMS calls
            is_ims_call = any(re.search(pattern, line) for pattern in ims_entry_points)
            
            # We no longer check for dynamic calls based just on USING parameters
            # as this leads to false positives. Only explicit IMS entry points are considered.
            
            # If it's an IMS call, process it
            if is_ims_call:
                call_lines = [lines[i].strip()]
                j = i + 1
                using_param = None
                local_parts = None

                if "USING" in line:
                    try:
                        local_parts = re.split(r"\bUSING\b", line, flags=re.IGNORECASE)[1].strip().split()
                        if local_parts:
                            using_param = local_parts[0]
                    except IndexError:
                        pass

                # Look for parameters on subsequent lines
                while j < n:
                    next_line = lines[j].strip()
                    call_lines.append(next_line)
                    if "USING" in next_line.upper():
                        try:
                            local_parts = re.split(r"\bUSING\b", next_line, flags=re.IGNORECASE)[1].strip().split()
                            if local_parts:
                                using_param = local_parts[0]
                        except IndexError:
                            pass
                        break
                    j += 1

                full_call = ' '.join(call_lines)
                
                # Determine if this is a static or dynamic call
                if any(re.search(pattern, call_lines[0]) for pattern in ims_entry_points):
                    call_type = "Static"
                else:
                    call_type = "Dynamic"
                
                # Extract function code if possible
                function_code = "Unknown"
                # Check if we have a parameter to analyze
                if using_param:
                    for func in ims_functions:
                        if func in using_param:
                            function_code = func
                            break
                
                ims_calls.append((program_name, 'COBOL', call_type, i+1, full_call, function_code))

                # Add parameter to counts if valid
                if using_param:
                    param = using_param.strip().replace('.', '')
                    if re.match(r'^[A-Z0-9\-]+$', param):
                        param_counts[param] = param_counts.get(param, 0) + 1

                i = j if j > i else i + 1
            else:
                i += 1

        else:

            i += 1

    return ims_calls, param_counts

def detect_ims_in_assembler(file_content: str, program_name: str):
    """Detect IMS calls in Assembler source code using universal patterns.
    
    Looks for standard IMS entry points and function codes to identify true IMS calls.
    """
    ims_calls = []
    lines = file_content.splitlines()
    
    # IMS entry point patterns for assembler
    ims_entry_points = [
        r'\bDFSLI00\d\b', r'\bAERTDLI\b', r'\bCBLTDLI\b', r'\bPLITDLI\b'
    ]
    
    # IMS function codes
    ims_functions = [
        'GU', 'GN', 'GNP', 'GHU', 'GHN', 'GHNP',
        'ISRT', 'DLET', 'REPL', 'CHKP', 'XRST', 'ROLL',
        'ROLS', 'SETS', 'SETU', 'INIT', 'INQY', 'CMD',
        'GCMD', 'AUTH', 'QUERY', 'POS', 'PCB'
    ]
    
    for i, line in enumerate(lines):
        upper = line.upper()
        
        # Check for IMS entry point references - only explicit IMS entry points
        is_ims_call = any(re.search(pattern, upper) for pattern in ims_entry_points)
        
        # We no longer check for function codes in generic instructions
        # as this leads to false positives
        
        # Check for CALL instruction with IMS entry points
        if re.search(r'\bCALL\b', upper):
            for pattern in ims_entry_points:
                if re.search(pattern, upper):
                    is_ims_call = True
                    break
        
        if is_ims_call:
            # Determine call type
            call_type = "Static"
            if "MVC" in upper and any(re.search(pattern, upper) for pattern in ims_entry_points):
                call_type = "Dynamic"
            
            # Try to identify function code
            function_code = "Unknown"
            for func in ims_functions:
                if func in upper:
                    function_code = func
                    break
            
            ims_calls.append((program_name, 'ASM', call_type, i+1, line.strip(), function_code))



    return ims_calls

def scan_all(cobol_folder, copybook_folder, asm_folder):
    """Scan all source files for IMS calls and generate comprehensive analysis."""
    cobol_files = []
    asm_files = []

    if cobol_folder and os.path.isdir(cobol_folder):
        cobol_files = [os.path.join(cobol_folder, f) for f in os.listdir(cobol_folder) 
                      if f.lower().endswith(('.cbl', '.cob', '.cobol'))]

    if asm_folder and os.path.isdir(asm_folder):
        asm_files = [os.path.join(asm_folder, f) for f in os.listdir(asm_folder) 
                    if f.lower().endswith(('.asm', '.s', '.mac'))]

    all_ims_calls = []
    all_param_counts = {}
    function_counts = {}
    cobol_with_ims = 0
    asm_with_ims = 0

    for file in cobol_files:
        try:
            # Try UTF-8 first
            with open(file, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
        except UnicodeDecodeError:
            # Fall back to Latin-1 which can read any byte
            with open(file, 'r', encoding='latin-1') as f:
                content = f.read()

        full_content = resolve_copybooks(content, copybook_folder)

        ims_calls, param_counts = detect_ims_in_cobol(full_content, os.path.basename(file))

        all_ims_calls.extend(ims_calls)
        
        # Count IMS function codes
        for _, _, _, _, _, function_code in ims_calls:
            function_counts[function_code] = function_counts.get(function_code, 0) + 1
        
        # Count parameters
        for param, count in param_counts.items():
            all_param_counts[param] = all_param_counts.get(param, 0) + count

        if ims_calls:
            cobol_with_ims += 1

    for file in asm_files:
        try:
            # Try UTF-8 first
            with open(file, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
        except UnicodeDecodeError:
            # Fall back to Latin-1 which can read any byte
            with open(file, 'r', encoding='latin-1') as f:
                content = f.read()

        ims_calls = detect_ims_in_assembler(content, os.path.basename(file))

        all_ims_calls.extend(ims_calls)

        if ims_calls:
            asm_with_ims += 1

    summary = {
        'Total COBOL Programs': len(cobol_files),
        'Total ASM Programs': len(asm_files),
        'COBOL with IMS Calls': cobol_with_ims,
        'ASM with IMS Calls': asm_with_ims,
        'COBOL without IMS Calls': len(cobol_files) - cobol_with_ims,
        'ASM without IMS Calls': len(asm_files) - asm_with_ims,
        'Total IMS-Using Programs': cobol_with_ims + asm_with_ims,
        'IMS Usage %': round((cobol_with_ims + asm_with_ims) / max(1, len(cobol_files) + len(asm_files)) * 100, 2)
    }

    return summary, all_ims_calls, all_param_counts, function_counts

def generate_clean_parameter_report_excel(parameter_map: dict, workbook: Workbook):

    ws = workbook.create_sheet("IMS Parameters Clean")

    ws.append(["Parameter", "Count"])

    for param, count in sorted(parameter_map.items(), key=lambda x: x[0]):

        ws.append([param, count])

    return workbook

def generate_final_excel_report_with_clean_params(summary: dict, inventory: list, clean_param_map: dict, function_counts: dict=None, output_dir=None):

    summary_df = pd.DataFrame(list(summary.items()), columns=["Metric", "Count"])

    # Handle both old format (5 elements) and new format (6 elements) calls
    inventory_with_functions = []
    for call in inventory:
        if len(call) == 5:
            inventory_with_functions.append(list(call) + ["Unknown"])
        else:
            inventory_with_functions.append(call)
            
    inventory_df = pd.DataFrame(inventory_with_functions, columns=["Program", "Language", "Static/Dynamic", "Line #", "Extracted IMS Call", "Function Code"])

    # Use output directory if provided
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        param_csv = os.path.join(output_dir, "ims_parameters.csv")
        summary_csv = os.path.join(output_dir, "ims_summary.csv")
        inventory_csv = os.path.join(output_dir, "ims_call_inventory.csv")
        function_csv = os.path.join(output_dir, "ims_functions.csv")
        excel_path = os.path.join(output_dir, "IMS_Impact_Report.xlsx")
    else:
        param_csv = "ims_parameters.csv"
        summary_csv = "ims_summary.csv"
        inventory_csv = "ims_call_inventory.csv"
        function_csv = "ims_functions.csv"
        excel_path = "IMS_Impact_Report.xlsx"

    pd.DataFrame(list(clean_param_map.items()), columns=["Parameter", "Count"]).to_csv(param_csv, index=False)
    summary_df.to_csv(summary_csv, index=False)
    inventory_df.to_csv(inventory_csv, index=False)
    if function_counts:
        pd.DataFrame(list(function_counts.items()), columns=["Function Code", "Count"]).to_csv(function_csv, index=False)

    wb = Workbook()

    ws1 = wb.active

    ws1.title = "Summary"

    for r in dataframe_to_rows(summary_df, index=False, header=True):

        ws1.append(r)

    ws2 = wb.create_sheet("IMS Call Inventory")

    for r in dataframe_to_rows(inventory_df, index=False, header=True):

        ws2.append(r)

    generate_clean_parameter_report_excel(clean_param_map, wb)

    # Save Excel file to output directory if provided
    wb.save(excel_path)

    print("\n" + "=" * 60)

    print("IMS Impact Analyzer Completed")

    # Display paths with output directory if provided
    if output_dir:
        print(f"Reports: {os.path.join(output_dir, 'ims_summary.csv')} | {os.path.join(output_dir, 'ims_call_inventory.csv')} | {os.path.join(output_dir, 'ims_parameters.csv')} | {os.path.join(output_dir, 'ims_functions.csv')}")
        print(f"Excel:  {os.path.join(output_dir, 'IMS_Impact_Report.xlsx')}")
    else:
        print("Reports: ims_summary.csv | ims_call_inventory.csv | ims_parameters.csv | ims_functions.csv")
        print("Excel:  IMS_Impact_Report.xlsx")

    print("\n" + "-" * 60)

def main(cobol_folder, copybook_folder, asm_folder, output_dir=None):
    """Main function to run the IMS analyzer with the improved detection logic."""
    summary, inventory, param_counts, function_counts = scan_all(cobol_folder, copybook_folder, asm_folder)

    generate_final_excel_report_with_clean_params(summary, inventory, param_counts, function_counts, output_dir)

# --- Main ---

if __name__ == "__main__":
    cobol_folder = input("Enter the path to the COBOL programs folder:\n> ").strip()

    copybook_folder = input("Enter the path to the COBOL Copybooks folder:\n> ").strip()

    asm_folder = input("Enter the path to the ASM (Assembler listing) folder:\n> ").strip()

    output_dir = input("Enter the output directory (optional):\n> ").strip()

    main(cobol_folder, copybook_folder, asm_folder, output_dir)