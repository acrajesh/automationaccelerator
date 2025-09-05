import zipfile
import os
import datetime
import pandas as pd
import sys

def scan_tools(tools_dir, report_path):
    try:
        try:
            now = datetime.datetime.now(datetime.timezone.utc)
            results = []
            for tool in sorted(os.listdir(tools_dir)):
                path = os.path.join(tools_dir, tool)
                # ignore non-directories (including .sh, .pl, .py files)
                if not os.path.isdir(path):
                    continue
                # determine JAR path, special case for turnstone
                if tool.lower() == 'turnstone':
                    bin_dir = os.path.join(path, 'binaries')
                    jar = os.path.join(bin_dir, f"{tool}.jar") if os.path.isdir(bin_dir) else os.path.join(path, f"{tool}.jar")
                else:
                    jar = os.path.join(path, f"{tool}.jar")
                version = "Not available"
                time_str = "N/A"
                age_str = "N/A"
                # extract manifest if JAR exists
                if os.path.isfile(jar):
                    try:
                        with zipfile.ZipFile(jar) as z:
                            with z.open("META-INF/MANIFEST.MF") as f:
                                content = f.read().decode("utf-8")
                        for line in content.splitlines():
                            if line.startswith("Implementation-Version:"):
                                version = line.split(":",1)[1].strip()
                            elif line.startswith("Build-Time:"):
                                time_str = line.split(":",1)[1].strip()
                        if version != "Not available" and time_str != "N/A":
                            bt = datetime.datetime.fromisoformat(time_str.replace("Z", "+00:00"))
                            age = now - bt
                            if age.days >= 1:
                                age_str = f"{age.days} days"
                            elif age.seconds >= 3600:
                                age_str = f"{age.seconds//3600} hours"
                            else:
                                age_str = f"{age.seconds//60} minutes"
                    except Exception:
                        pass
                else:
                    # check for executable
                    exe = os.path.join(path, f"{tool}.exe")
                    if os.path.isfile(exe):
                        version = "Executable - please contact tool product team"
                        time_str = ""  # Empty string instead of N/A for executables
                        age_str = ""   # Empty string instead of N/A for executables
                results.append({
                    "Tool": tool,
                    "Implementation-Version": version,
                    "Build-Time": time_str,
                    "Age": age_str
                })
            df = pd.DataFrame(results)
            # Filter out "Not available" tools for the report
            df_filtered = df[df['Implementation-Version'] != "Not available"]
            
            # Split into categories
            tools_with_version = df_filtered[df_filtered['Implementation-Version'] != "Executable - please contact tool product team"]
            executable_tools = df_filtered[df_filtered['Implementation-Version'] == "Executable - please contact tool product team"]
            
            # Attempt formatted Excel output; fall back to CSV if openpyxl is missing
            try:
                from openpyxl.styles import Alignment, Font, PatternFill
                from openpyxl.utils import get_column_letter
                
                with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                    # Create a new DataFrame with a category column for sorting
                    df_report = pd.DataFrame()
                    
                    # Add section headers and tools with versions
                    if not tools_with_version.empty:
                        header_row = pd.DataFrame([{"Tool": "=== Tools with Version Information ===", 
                                                  "Implementation-Version": "", 
                                                  "Build-Time": "", 
                                                  "Age": ""}])
                        df_report = pd.concat([df_report, header_row, tools_with_version])
                    
                    # Add executable tools section
                    if not executable_tools.empty:
                        # Add an empty row for spacing
                        empty_row = pd.DataFrame([{"Tool": "", "Implementation-Version": "", "Build-Time": "", "Age": ""}])
                        header_row = pd.DataFrame([{"Tool": "=== Executable Tools - Contact Product Team ===", 
                                                  "Implementation-Version": "", 
                                                  "Build-Time": "", 
                                                  "Age": ""}])
                        df_report = pd.concat([df_report, empty_row, header_row, executable_tools])
                    
                    # Write to Excel
                    df_report.to_excel(writer, index=False, sheet_name='Report')
                    ws = writer.sheets['Report']
                    
                    # Format header row
                    header_font = Font(bold=True)
                    for cell in ws[1]:
                        cell.font = header_font
                    
                    # Format section headers
                    section_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    for row in ws.iter_rows():
                        if row[0].value and str(row[0].value).startswith("==="):
                            for cell in row:
                                cell.font = header_font
                                cell.fill = section_fill
                    
                    # Set column widths and cell formatting
                    for i, col in enumerate(ws.columns):
                        max_len = max(len(str(cell.value) or "") for cell in col)
                        col_letter = get_column_letter(i+1)
                        ws.column_dimensions[col_letter].width = max_len + 4  # Add more padding
                        for cell in col:
                            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
                    
                    # Set row heights for better spacing
                    for i in range(2, ws.max_row + 1):
                        ws.row_dimensions[i].height = 25  # Increase row height for better spacing
                        if ws.cell(row=i, column=1).value and str(ws.cell(row=i, column=1).value).startswith("==="):
                            ws.row_dimensions[i].height = 30  # Increase row height for section headers
                            if i+1 <= ws.max_row:
                                ws.row_dimensions[i+1].height = 20  # Add some space after section headers
                            if i-1 >= 1:
                                ws.row_dimensions[i-1].height = 20  # Add some space before section headers
                # Don't print Excel report saved message here
            except ImportError:
                csv_path = report_path.replace('.xlsx', '.csv')
                print("openpyxl not installed; writing CSV report instead.")
                df.to_csv(csv_path, index=False)
                print(f"CSV report saved to {csv_path}")
            return df
        except Exception as e:
            print("Exception in scan_tools:", e)
            raise
    except Exception as e:
        print("Exception in scan_tools:", e)
        raise

def display_selected_tools(df):
    """Display a formatted table of all tools in three sections"""
    if not df.empty:
        # Split the dataframe into three categories
        tools_with_version = df[(df['Implementation-Version'] != "Not available") & 
                               (df['Implementation-Version'] != "Executable - please contact tool product team")]
        
        executable_tools = df[df['Implementation-Version'] == "Executable - please contact tool product team"]
        
        # Calculate the maximum width needed for each column
        tool_width = max(12, df['Tool'].str.len().max())
        version_width = max(30, df['Implementation-Version'].str.len().max())
        build_time_width = 25  # Keep this fixed as it's a standard format
        age_width = 10  # Keep this fixed as it's usually short
        
        # 1. Display tools with version information
        if not tools_with_version.empty:
            print(f"\n=== Tools with Version Information ===\n")
            print(f"{'Tool':<{tool_width}} {'Version':<{version_width}} {'Build-Time':<{build_time_width}} {'Age':<{age_width}}")
            print(f"{'----':<{tool_width}} {'---------':<{version_width}} {'---------':<{build_time_width}} {'---':<{age_width}}")
            
            for _, row in tools_with_version.iterrows():
                print(f"{row['Tool']:<{tool_width}} {row['Implementation-Version']:<{version_width}} {row['Build-Time']:<{build_time_width}} {row['Age']:<{age_width}}")
        
        # 2. Display executable tools
        if not executable_tools.empty:
            print(f"\n=== Executable Tools - Contact Product Team ===\n")
            print(f"{'Tool':<{tool_width}}")
            print(f"{'----':<{tool_width}}")
            
            for _, row in executable_tools.iterrows():
                print(f"{row['Tool']:<{tool_width}}")
                
        return True
    else:
        print('No tools found')
        return False

# Main function is removed as it will be called from the main CLI script
