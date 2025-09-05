import re
import sys
import os
import pandas as pd
from openpyxl import Workbook

def extract_summary_table(log_lines):
    summary_tables = []
    capture = False
    table_data = []
    column_count = 9  # Adjusted to exclude 'Generated' column
    
    for line in log_lines:
        if "Summary:" in line:
            capture = True
            continue
        
        if capture:
            if re.search(r'\+[-]+', line):  # Detect table start or end
                if table_data:
                    summary_tables.append(table_data)
                    table_data = []
            elif "|" in line and "Parse error" not in line:
                clean_line = re.sub(r'^[^|]+\|', '', line).strip()  # Remove log prefixes
                parts = [x.strip() for x in clean_line.split('|') if x.strip()]
                
                if len(parts) >= column_count:  # Ensure all columns exist
                    table_data.append(parts[:column_count])  # Capture first 9 columns
    
    return summary_tables

def extract_missing_copybooks(log_lines):
    missing_copybooks = set()
    # Updated pattern to handle different path formats
    pattern = re.compile(r'\[(?:.*[\\/])?([^\]\/:]+)\.cbl(?:[:L\d+]*)?\] .* Copybook error: Copybook does not exist ([\w$@#]+)')
    
    for line in log_lines:
        match = pattern.search(line)
        if match:
            program, copybook = match.groups()
            missing_copybooks.add((program, copybook))  # Use set to remove duplicates
    
    return list(missing_copybooks)

def extract_parse_errors(log_lines):
    parse_errors = []
    # Updated pattern to handle different path formats
    pattern = re.compile(r'\[(?:.*[\\/])?([^\]\/:]+)\.cbl\].*?Parse error .*?\[(.*?)\]')
    
    for line in log_lines:
        match = pattern.search(line)
        if match:
            program, error_message = match.groups()
            parse_errors.append([program, error_message])
    
    return parse_errors

def create_excel_report(log_file_path, output_dir=None):
    app_name = input("Enter App Name for the Excel file: ")
    
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        excel_file = os.path.join(output_dir, f"{app_name}.xlsx")
    else:
        excel_file = f"{app_name}.xlsx"
    
    with open(log_file_path, 'r', encoding='utf-8') as file:
        log_lines = file.readlines()
    
    summary_tables = extract_summary_table(log_lines)
    missing_copybooks = extract_missing_copybooks(log_lines)
    parse_errors = extract_parse_errors(log_lines)
    
    wb = Workbook()
    
    # Conversion Status Sheet
    ws1 = wb.active
    ws1.title = "Conversion Status"
    ws1.append(["Filename", "Errors-Total", "Errors-SQL", "Errors-CICS", "Errors-DLI", "Warnings-Total", "Warnings-SQL", "Warnings-CICS", "Warnings-DLI"])
    
    for table in summary_tables:
        for row in table:
            if "Parse error" not in row[0]:  # Ensure parse errors are not mistakenly added
                ws1.append(row)
    
    # Missing Copybooks Sheet
    ws2 = wb.create_sheet(title="Missing Copybooks")
    ws2.append(["Program", "Missing Copybook"])
    
    for row in missing_copybooks:
        ws2.append(row)
    
    # Full Log Sheet
    ws3 = wb.create_sheet(title="log-file")
    for line in log_lines:
        ws3.append([line.strip()])
    
    # COBOL Status Sheet
    ws4 = wb.create_sheet(title="COBOL Status")
    ws4.append(["Type", "Total", "Done", "Done Percentage", "Error", "Error Percentage"])
    
    if summary_tables:
        total_files = sum(len(table) for table in summary_tables)
        done_count = sum(1 for table in summary_tables for row in table if len(row) > 1 and row[1].isdigit() and int(row[1]) == 0)
    else:
        total_files = 0
        done_count = 0
    error_count = total_files - done_count
    done_percentage = round((done_count / total_files) * 100, 0) if total_files else 0
    error_percentage = round((error_count / total_files) * 100, 0) if total_files else 0
    
    ws4.append(["Cobol Programs", total_files, done_count, f"{done_percentage}%", error_count, f"{error_percentage}%"])
    
    # Parse Errors Sheet
    ws5 = wb.create_sheet(title="Parse Errors")
    ws5.append(["Program", "Error Message"])
    
    for row in parse_errors:
        ws5.append(row)
    
    wb.save(excel_file)
    print(f"Excel report saved as {excel_file}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python cobol_convertlog_scanner.py logfile.log [--output-dir OUTPUT_DIR]")
        sys.exit(1)
    
    log_file_path = sys.argv[1]
    output_dir = None
    
    # Check for output directory argument
    if len(sys.argv) > 2 and sys.argv[2] == "--output-dir" and len(sys.argv) > 3:
        output_dir = sys.argv[3]
        
    create_excel_report(log_file_path, output_dir)
